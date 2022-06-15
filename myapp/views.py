from django.shortcuts import render
import openpyxl
from django.http import HttpResponse
from django.views.generic import ListView

def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)

        # getting a particular sheet
        worksheet = wb[sheets[0]]
        #print(worksheet)


        # getting active sheet
        #active_sheet = wb.active
        #print(active_sheet)

        # reading a cell
        #print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            temp = 0
            for cell in row:
                
                value = ""
                if cell.value is not None:
                    value = str(cell.value)
                    if temp == 2:
                      value = value + 'second' 
                      ### CHECK VALUE INSIDE DATABASE ####################
                      # if in database:
                      
                      # temp test REPLACE LATER WITH 'IF IN DATABASE:'
                      if "鋼" in value:
                          value = value + '!'
                temp += 1
                row_data.append(value)
                #print(cell.value)
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data":excel_data})


def search(request):
    if "GET" == request.method:
        return render(request, 'myapp/results.html', {})
    else:
        result_list = list()
           
        form = request.POST
        #print(form['results_btn'])
        
        keyword_str = form['results_btn']
        print(keyword_str)

        #test!!!!!!!!!!!!
        result_list.append('apple')
        result_list.append('book')
        result_list.append('choo choo train')
        result_list.append(keyword_str)


        return render(request, 'myapp/results.html', {"result_list":result_list})

 
